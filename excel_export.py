"""
STAATS Excel Exporter
Creates professionally formatted Excel output matching STAATS style

Features:
- Multiple tabs per file
- Conditional formatting (significance highlighting)
- Summary/index sheet
- Proper number formatting
- Cell styling (headers, percentages, counts)
- Auto-column sizing
"""

from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

from tab_engine import TabResult, TabEngine
from core import TabSpec, DataMap


class ExcelFormatter:
    """
    Styling constants and helpers
    """
    # Colors (matching professional research output)
    HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    SUBHEADER_FILL = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    TOTAL_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    SIGNIFICANT_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    # Fonts
    HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    TITLE_FONT = Font(name='Arial', size=12, bold=True)
    NORMAL_FONT = Font(name='Arial', size=10)
    TOTAL_FONT = Font(name='Arial', size=10, bold=True)
    
    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alignment
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT = Alignment(horizontal='left', vertical='center')
    RIGHT = Alignment(horizontal='right', vertical='center')
    
    @staticmethod
    def auto_column_width(worksheet, min_width=8, max_width=50):
        """Auto-adjust column widths based on content"""
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            adjusted_width = min(max(length + 2, min_width), max_width)
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width


class ExcelExporter:
    """
    Export TabResults to Excel with professional formatting
    """
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def export_single_tab(
        self,
        result: TabResult,
        filepath: str,
        sheet_name: str = "Tab",
        display_mode: str = "Both"
    ):
        """
        Export a single tab result to Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel limit
        
        self._write_tab_to_sheet(ws, result, display_mode)
        
        wb.save(filepath)
    
    def export_multiple_tabs(
        self,
        results: List[TabResult],
        filename: str,
        display_mode: str = "Both",
        create_summary: bool = True
    ):
        """
        Export multiple tabs to a single Excel file
        Each tab gets its own sheet
        Optional summary/index sheet
        """
        filepath = self.output_dir / filename
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create summary sheet if requested
        if create_summary:
            summary_ws = wb.create_sheet("Summary", 0)
            self._create_summary_sheet(summary_ws, results)
        
        # Create sheet for each tab
        for i, result in enumerate(results, 1):
            # Sanitize sheet name (max 31 chars, no special chars)
            sheet_name = result.title[:28]
            sheet_name = ''.join(c for c in sheet_name if c.isalnum() or c in ' -_')
            if not sheet_name:
                sheet_name = f"Tab{i}"
            
            # Ensure unique name
            if sheet_name in wb.sheetnames:
                sheet_name = f"{sheet_name}_{i}"
            
            ws = wb.create_sheet(sheet_name)
            self._write_tab_to_sheet(ws, result, display_mode)
        
        wb.save(filepath)
        print(f"✅ Exported {len(results)} tabs to {filepath}")
        
        return filepath
    
    def _write_tab_to_sheet(
        self,
        ws,
        result: TabResult,
        display_mode: str = "Both"
    ):
        """
        Write a TabResult to a worksheet with formatting
        """
        # Title
        ws['A1'] = result.title
        ws['A1'].font = ExcelFormatter.TITLE_FONT
        
        # Get display data
        if display_mode == "Vertical":
            display_df = result.col_pct
            format_type = 'percentage'
        elif display_mode == "Horizontal":
            display_df = result.row_pct
            format_type = 'percentage'
        else:  # Both
            display_df = result.to_display_format("Both")
            format_type = 'combined'
        
        if display_df.empty:
            ws['A3'] = "No data"
            return
        
        # Base counts
        if result.base is not None:
            row = 3
            ws[f'A{row}'] = 'Base (n):'
            ws[f'A{row}'].font = ExcelFormatter.TOTAL_FONT
            
            for col_idx, (col_name, base_val) in enumerate(result.base.items(), 2):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = int(base_val) if pd.notna(base_val) else '-'
                cell.font = ExcelFormatter.TOTAL_FONT
                cell.alignment = ExcelFormatter.CENTER
        
        # Data starts at row 5
        start_row = 5
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(display_df, index=True, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                
                # Format based on position
                if r_idx == start_row:  # Header row
                    cell.value = value
                    cell.font = ExcelFormatter.HEADER_FONT
                    cell.fill = ExcelFormatter.HEADER_FILL
                    cell.alignment = ExcelFormatter.CENTER
                    cell.border = ExcelFormatter.THIN_BORDER
                
                elif c_idx == 1:  # Row labels
                    cell.value = value
                    if value == 'Total':
                        cell.font = ExcelFormatter.TOTAL_FONT
                        cell.fill = ExcelFormatter.TOTAL_FILL
                    else:
                        cell.font = ExcelFormatter.NORMAL_FONT
                    cell.alignment = ExcelFormatter.LEFT
                    cell.border = ExcelFormatter.THIN_BORDER
                
                else:  # Data cells
                    # Parse value based on format
                    if format_type == 'percentage':
                        if isinstance(value, (int, float)):
                            cell.value = value / 100
                            cell.number_format = '0.0%'
                        else:
                            cell.value = value
                    elif format_type == 'combined':
                        cell.value = value
                    else:
                        cell.value = value
                    
                    # Styling - check if this is a Total row
                    row_label_cell = ws.cell(row=r_idx, column=1)
                    if row_label_cell.value == 'Total':
                        cell.font = ExcelFormatter.TOTAL_FONT
                        cell.fill = ExcelFormatter.TOTAL_FILL
                    else:
                        cell.font = ExcelFormatter.NORMAL_FONT
                    
                    cell.alignment = ExcelFormatter.CENTER
                    cell.border = ExcelFormatter.THIN_BORDER
        
        # Add significance markers if available
        if result.significance is not None and not result.significance.empty:
            sig_start_row = start_row + len(display_df) + 3
            
            ws[f'A{sig_start_row}'] = 'Significance (columns significantly higher):'
            ws[f'A{sig_start_row}'].font = Font(italic=True, size=9)
            
            for r_idx, row in enumerate(dataframe_to_rows(result.significance, index=True, header=True), sig_start_row + 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    cell.font = Font(size=9)
                    cell.alignment = ExcelFormatter.CENTER
                    
                    # Highlight cells with significance
                    if c_idx > 1 and value and str(value).strip():
                        cell.fill = ExcelFormatter.SIGNIFICANT_FILL
        
        # Note
        note_row = start_row + len(display_df) + 10
        if result.weighted:
            ws[f'A{note_row}'] = 'Note: Results are weighted'
            ws[f'A{note_row}'].font = Font(italic=True, size=9)
        
        # Auto-size columns
        ExcelFormatter.auto_column_width(ws)
    
    def _create_summary_sheet(self, ws, results: List[TabResult]):
        """
        Create index/summary sheet listing all tabs
        """
        ws['A1'] = 'Table of Contents'
        ws['A1'].font = ExcelFormatter.TITLE_FONT
        
        ws['A3'] = 'Tab'
        ws['B3'] = 'Title'
        ws['C3'] = 'Base'
        
        for col in ['A3', 'B3', 'C3']:
            ws[col].font = ExcelFormatter.HEADER_FONT
            ws[col].fill = ExcelFormatter.HEADER_FILL
        
        for i, result in enumerate(results, 4):
            ws[f'A{i}'] = i - 3
            ws[f'B{i}'] = result.title
            
            if result.base is not None and 'Total' in result.base:
                ws[f'C{i}'] = int(result.base['Total'])
            else:
                ws[f'C{i}'] = '-'
        
        ExcelFormatter.auto_column_width(ws)


# Example usage
if __name__ == "__main__":
    from tab_engine import TabEngine, TabResult
    from core import DataMap, Question, QuestionType, TabDefinition, Filter, Class
    from engines import FilterEngine, ClassEngine
    import pandas as pd
    import numpy as np
    
    print("🧪 Testing Excel Exporter\n")
    
    # Create test data
    np.random.seed(42)
    n = 200
    
    df = pd.DataFrame({
        'Country': np.random.choice([1, 2, 3], n, p=[0.5, 0.3, 0.2]),
        'Age': np.random.randint(25, 65, n),
        'Satisfaction': np.random.choice([1, 2, 3, 4, 5], n),
        'Recommend': np.random.choice([1, 2, 3], n),
    })
    
    # Setup datamap
    dm = DataMap()
    dm.add_question(Question('Country', QuestionType.QUALI_UNIQUE, 'Country', {
        1: 'France', 2: 'UK', 3: 'Germany'
    }))
    dm.add_question(Question('Age', QuestionType.NUMERIC, 'Age'))
    dm.add_question(Question('Satisfaction', QuestionType.QUALI_UNIQUE, 'Satisfaction', {
        1: 'Very dissatisfied', 2: 'Dissatisfied', 3: 'Neutral',
        4: 'Satisfied', 5: 'Very satisfied'
    }))
    dm.add_question(Question('Recommend', QuestionType.QUALI_UNIQUE, 'Recommend', {
        1: 'Yes', 2: 'No', 3: 'Maybe'
    }))
    
    # Setup class
    ce = ClassEngine()
    ce.add_class(Class('AgeGroups', [
        ('X>=25 and X<40', 'Under 40'),
        ('X>=40 and X<55', '40-54'),
        ('X>=55', '55+')
    ]))
    
    # Create tab engine
    tab_engine = TabEngine(dm, class_engine=ce)
    
    # Generate multiple tabs
    specs = [
        TabDefinition('Satisfaction by Country', 'Satisfaction', 'Country'),
        TabDefinition('Recommend by Country', 'Recommend', 'Country'),
        TabDefinition('Age Groups by Country', 'Age', 'Country', class_name='AgeGroups'),
        TabDefinition('Satisfaction by Recommend', 'Satisfaction', 'Recommend'),
    ]
    
    results = tab_engine.generate_multiple_tabs(df, specs)
    
    # Export to Excel
    exporter = ExcelExporter(output_dir="/mnt/user-data/outputs")
    
    print("Exporting tabs to Excel...")
    filepath = exporter.export_multiple_tabs(
        results,
        filename='survey_analysis.xlsx',
        display_mode='Both',
        create_summary=True
    )
    
    print(f"\n✅ Excel export complete!")
    print(f"File: {filepath}")
    print(f"Tabs: {len(results)}")
