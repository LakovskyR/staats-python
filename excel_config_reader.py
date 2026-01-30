"""
STAATS Excel Config Reader
Parse existing STAATS.xlsm files and convert to Python objects

Reads from:
- Datamap tab → DataMap object
- Recode tab → RecodeEngine with all recodes
- Filters tab → FilterEngine
- Classes tab → ClassEngine
- Tab specifications tab → List of TabSpec objects
"""

from typing import Dict, List, Optional, Tuple
import pandas as pd
import openpyxl
from pathlib import Path

from core import (
    DataMap, Question, QuestionType, Filter, Class,
    TabDefinition, TabSpec, RecodeType
)
from recode_engine import (
    RecodeEngine, QualiUniqueRecode, QualiMultipleRecode,
    NumericRecode, NumberOfAnswersRecode, CombinationRecode,
    WeightRecode, QualiMultiINIRecode
)
from engines import FilterEngine, ClassEngine


class ExcelConfigReader:
    """
    Read STAATS configuration from Excel files
    Maintains backward compatibility with existing STAATS.xlsm
    """
    
    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            raise FileNotFoundError(f"Config file not found: {filepath}")
        
        # Load workbook
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
    
    def read_datamap(self) -> DataMap:
        """
        Read Datamap tab and create DataMap object
        
        Expected columns:
        - Name: Variable name
        - Type: Question type (QU, QM, N, O)
        - Title: Question label
        - Code columns: Code table (varies by question)
        """
        if 'Datamap' not in self.wb.sheetnames:
            raise ValueError("Datamap sheet not found in workbook")
        
        ws = self.wb['Datamap']
        
        dm = DataMap()
        
        # Parse header to find column positions
        header_row = None
        for row_idx in range(1, 20):
            row_values = [ws.cell(row_idx, col).value for col in range(1, 10)]
            if 'Name' in row_values and 'Type' in row_values:
                header_row = row_idx
                break
        
        if not header_row:
            print("Warning: Could not find Datamap header row")
            return dm
        
        # Get column positions
        col_positions = {}
        for col_idx in range(1, 20):
            value = ws.cell(header_row, col_idx).value
            if value:
                col_positions[value] = col_idx
        
        if 'Name' not in col_positions or 'Type' not in col_positions:
            print("Warning: Name or Type column not found in Datamap")
            return dm
        
        # Read questions
        for row_idx in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row_idx, col_positions['Name']).value
            qtype_str = ws.cell(row_idx, col_positions['Type']).value
            
            if not name or not qtype_str:
                continue
            
            # Parse question type
            try:
                qtype = QuestionType.from_string(qtype_str)
            except:
                qtype = QuestionType.OPEN
            
            # Get title
            title = ws.cell(row_idx, col_positions.get('Title', col_positions['Name'] + 1)).value or name
            
            # Parse code table (usually starts after Title column)
            codes = {}
            code_start_col = col_positions.get('Title', col_positions['Type'] + 1) + 1
            
            for col_offset in range(0, 50, 2):  # Codes are in pairs: code, label
                code_col = code_start_col + col_offset
                label_col = code_start_col + col_offset + 1
                
                if code_col > ws.max_column:
                    break
                
                code_val = ws.cell(row_idx, code_col).value
                label_val = ws.cell(row_idx, label_col).value
                
                if code_val is not None and label_val is not None:
                    try:
                        codes[int(code_val)] = str(label_val)
                    except (ValueError, TypeError):
                        pass
            
            question = Question(
                name=str(name),
                qtype=qtype,
                title=str(title),
                codes=codes
            )
            
            dm.add_question(question)
        
        return dm
    
    def read_recodes(self, datamap: DataMap) -> RecodeEngine:
        """
        Read Recode tab and create RecodeEngine
        
        Expected columns:
        - Name: Recode variable name
        - Type: Recode type
        - Title: Recode label
        - Option NA: Yes/No
        - Formula: Recode formula
        - Codes/Labels: Code table (for quali recodes)
        """
        if 'Recode' not in self.wb.sheetnames:
            print("Warning: Recode sheet not found")
            return RecodeEngine()
        
        ws = self.wb['Recode']
        engine = RecodeEngine()
        
        # Find header
        header_row = None
        for row_idx in range(1, 20):
            row_values = [ws.cell(row_idx, col).value for col in range(1, 15)]
            if 'Name' in row_values and 'Type' in row_values:
                header_row = row_idx
                break
        
        if not header_row:
            print("Warning: Could not find Recode header")
            return engine
        
        # Get column positions
        col_pos = {}
        for col_idx in range(1, 20):
            val = ws.cell(header_row, col_idx).value
            if val:
                col_pos[val] = col_idx
        
        # Read recodes
        current_recode = None
        formula_lines = []
        codes = {}
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row_idx, col_pos.get('Name', 1)).value
            
            # If we hit a new recode, save the previous one
            if name and current_recode:
                # Create recode object based on type
                try:
                    recode_obj = self._create_recode(
                        current_recode['name'],
                        current_recode['type'],
                        current_recode['title'],
                        '\n'.join(formula_lines),
                        codes,
                        current_recode['option_na']
                    )
                    engine.add_recode(recode_obj)
                except Exception as e:
                    print(f"Warning: Failed to create recode '{current_recode['name']}': {e}")
                
                # Reset for next recode
                formula_lines = []
                codes = {}
            
            # Start new recode
            if name:
                rtype = ws.cell(row_idx, col_pos.get('Type', 2)).value
                title = ws.cell(row_idx, col_pos.get('Title', 3)).value or name
                option_na = ws.cell(row_idx, col_pos.get('Option NA', 4)).value
                
                current_recode = {
                    'name': str(name),
                    'type': str(rtype) if rtype else 'quali_unique',
                    'title': str(title),
                    'option_na': str(option_na).lower() == 'yes' if option_na else False
                }
            
            # Collect formula lines and codes
            formula_cell = ws.cell(row_idx, col_pos.get('Formula', 5)).value
            if formula_cell:
                formula_lines.append(str(formula_cell))
            
            # Collect codes (usually after formula column)
            code_col = col_pos.get('Formula', 5) + 1
            label_col = code_col + 1
            
            code_val = ws.cell(row_idx, code_col).value
            label_val = ws.cell(row_idx, label_col).value
            
            if code_val is not None and label_val is not None:
                try:
                    codes[int(code_val)] = str(label_val)
                except (ValueError, TypeError):
                    pass
        
        # Don't forget the last recode
        if current_recode:
            try:
                recode_obj = self._create_recode(
                    current_recode['name'],
                    current_recode['type'],
                    current_recode['title'],
                    '\n'.join(formula_lines),
                    codes,
                    current_recode['option_na']
                )
                engine.add_recode(recode_obj)
            except Exception as e:
                print(f"Warning: Failed to create recode '{current_recode['name']}': {e}")
        
        return engine
    
    def _create_recode(self, name: str, rtype: str, title: str, formula: str, 
                       codes: Dict, option_na: bool):
        """Helper to create appropriate recode object"""
        rtype_lower = rtype.lower().replace(' ', '_')
        
        if 'quali_unique' in rtype_lower or 'qualitative_unique' in rtype_lower:
            return QualiUniqueRecode(name, title, formula, codes, option_na)
        
        elif 'quali_multi' in rtype_lower and 'ini' not in rtype_lower:
            return QualiMultipleRecode(name, title, formula, codes, option_na)
        
        elif 'numeric' in rtype_lower:
            return NumericRecode(name, RecodeType.NUMERIC, title, formula, option_na)
        
        elif 'number_of_answer' in rtype_lower or 'count' in rtype_lower:
            return NumberOfAnswersRecode(name, title, formula, option_na)
        
        elif 'combination' in rtype_lower:
            return CombinationRecode(name, title, formula, option_na)
        
        elif 'weight' in rtype_lower:
            # Parse weights from formula
            weights = {}
            for line in formula.split('\n'):
                if ':' in line:
                    cond, weight_str = line.split(':', 1)
                    try:
                        weights[cond.strip()] = float(weight_str.strip())
                    except:
                        pass
            return WeightRecode(name, title, '', weights, option_na)
        
        else:
            # Default to quali unique
            return QualiUniqueRecode(name, title, formula, codes, option_na)
    
    def read_filters(self) -> FilterEngine:
        """
        Read Filters tab
        
        Expected columns:
        - Name
        - Formula
        - With NA
        """
        if 'Filters' not in self.wb.sheetnames:
            print("Warning: Filters sheet not found")
            return FilterEngine()
        
        ws = self.wb['Filters']
        engine = FilterEngine()
        
        # Find header (usually row 1)
        for row_idx in range(1, ws.max_row + 1):
            name = ws.cell(row_idx, 2).value  # Column B
            formula = ws.cell(row_idx, 3).value  # Column C
            with_na = ws.cell(row_idx, 4).value  # Column D
            
            if name and formula and str(name) != 'Name':
                filter_obj = Filter(
                    name=str(name),
                    formula=str(formula),
                    with_na=str(with_na).lower() == 'yes' if with_na else False
                )
                engine.add_filter(filter_obj)
        
        return engine
    
    def read_classes(self) -> ClassEngine:
        """
        Read Classes tab
        
        Format: Each class is 2 columns (Formula, Label) with name in first row
        """
        if 'Classes' not in self.wb.sheetnames:
            print("Warning: Classes sheet not found")
            return ClassEngine()
        
        ws = self.wb['Classes']
        engine = ClassEngine()
        
        # Classes are in pairs of columns
        col = 2  # Start at column B
        
        while col < ws.max_column:
            # Get class name from row 1
            class_name = ws.cell(1, col).value
            
            if not class_name:
                col += 2
                continue
            
            # Get option NA from row 2
            option_na_val = ws.cell(2, col + 1).value
            option_na = str(option_na_val).lower() == 'yes' if option_na_val else False
            
            # Read bins (formula, label) pairs starting from row 4
            bins = []
            for row_idx in range(4, ws.max_row + 1):
                formula = ws.cell(row_idx, col).value
                label = ws.cell(row_idx, col + 1).value
                
                if formula and label:
                    bins.append((str(formula), str(label)))
                elif not formula:
                    break
            
            if bins:
                class_obj = Class(
                    name=str(class_name),
                    bins=bins,
                    option_na=option_na
                )
                engine.add_class(class_obj)
            
            col += 2  # Move to next class
        
        return engine
    
    def read_all(self) -> Tuple[DataMap, RecodeEngine, FilterEngine, ClassEngine]:
        """
        Read all configuration from Excel
        Returns: (datamap, recode_engine, filter_engine, class_engine)
        """
        print("📖 Reading STAATS configuration from Excel...")
        
        datamap = self.read_datamap()
        print(f"   ✓ Datamap: {len(datamap)} questions")
        
        recode_engine = self.read_recodes(datamap)
        print(f"   ✓ Recodes: {len(recode_engine)} recodes")
        
        filter_engine = self.read_filters()
        print(f"   ✓ Filters: {len(filter_engine)} filters")
        
        class_engine = self.read_classes()
        print(f"   ✓ Classes: {len(class_engine)} classes")
        
        return datamap, recode_engine, filter_engine, class_engine


# Example usage
if __name__ == "__main__":
    print("🧪 Testing Excel Config Reader\n")
    
    # Test with the project STAATS file
    config_path = "/mnt/project/STAATS_2__v2_25.xlsm"
    
    try:
        reader = ExcelConfigReader(config_path)
        
        datamap, recode_engine, filter_engine, class_engine = reader.read_all()
        
        print("\n" + "=" * 80)
        print("RESULTS")
        print("=" * 80)
        
        print(f"\n📊 Datamap: {len(datamap)} questions")
        if len(datamap) > 0:
            print("   Sample questions:")
            for i, (name, q) in enumerate(list(datamap.questions.items())[:5]):
                print(f"     - {name}: {q.qtype.name} - {q.title}")
        
        print(f"\n🔄 Recodes: {len(recode_engine)}")
        if len(recode_engine) > 0:
            print("   Sample recodes:")
            for recode in recode_engine.recodes[:5]:
                print(f"     - {recode.name}: {recode.rtype.name}")
        
        print(f"\n🎯 Filters: {len(filter_engine)}")
        if len(filter_engine) > 0:
            print("   Sample filters:")
            for name, filt in list(filter_engine.filters.items())[:5]:
                print(f"     - {name}: {filt.formula}")
        
        print(f"\n📐 Classes: {len(class_engine)}")
        if len(class_engine) > 0:
            print("   Sample classes:")
            for name, cls in list(class_engine.classes.items())[:5]:
                print(f"     - {name}: {len(cls.bins)} bins")
        
        print("\n✅ Excel config read complete!")
        
    except Exception as e:
        print(f"❌ Error reading config: {e}")
        import traceback
        traceback.print_exc()
